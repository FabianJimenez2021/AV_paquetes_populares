import openpyxl 

wb = openpyxl.load_workbook("Nombre_archivo.xlsx")

#sheetnames = devuelve los nombres de las hojas del archivo  
print(wb.sheetnames)

#Ver una hoja en específico 
print(wb["Hoja1"])

#active = devuelve la hoja que se encuentra activa 
hoja = wb.active
print(hoja)

#Create_sheet = Crea hojas de calculo 
wb.create_sheet("Hoja 3")

#Cambiar el titulo de una hoja de calculo 
hoja3 = wb["Hoja 3"]
hoja3.title = "Nuevo titulo"

#Imprime las filas y las columnas que están en la hoja de calculo (en este caso la hoja activa) 
print(
    hoja.max_row, 
    hoja.max_column
)

#Ver el valor que tiene una celda 
celda = hoja["A1"]
print(celda.value)

#Cambiar el valor de una celda
celda.value = "Nombre completo"

#Otra forma de ver el valor de las celdas 
#Nota: las columnas y las filas se cuentan a partir del indice 1 no de 0 
celda2 = hoja.cell(row=2, column=1)
print (
    celda2.value,
    celda2.row,
    celda2.column, 
    celda2.coordinate
)


#Ver todos los valores de la hoja de calculo 
for fila in range(1,hoja.max_row + 1):
    for columna in range(1, hoja.max_column + 1):
        celda = hoja.cell(row=fila, column=columna)
        print(fila, columna, celda.value)

#Ver los valores de toda la columna 
columna = hoja["A"]
print(columna)

#Ver los valores de toda la fila
fila = hoja["1"]
print(fila)

#Agregar valores a las filas 
hoja.append([1,2,3])
print(hoja.rows)

#Eliminar filas o columnas 
hoja.delete_rows(1, 1)

#Guardar el archivo 
wb.save("ruta_archivo")